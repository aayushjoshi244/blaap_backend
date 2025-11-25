import sys
import json
import time
import math
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set

import cv2
import numpy as np
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from datetime import datetime




# Excel logging (attendance)
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


# =========================
# Paths & constants
# =========================

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
EMB_DIR = DATA_DIR / "embeddings"
STUDENTS_JSON = DATA_DIR / "students.json"
RAW_DIR = DATA_DIR / "raw"
ATTENDANCE_FILE = DATA_DIR / "attendance.xlsx"
ATTENDANCE_BASE = DATA_DIR / "attendance" 


POSES = ["FRONT", "LEFT", "RIGHT"]
IMAGES_PER_POSE = 10          # fixed 10 per pose
SIM_THRESHOLD = 0.45          # cosine similarity threshold
CAM_INDEX = 0                 # USB camera index

APP_MODEL = None        # InsightFace FaceAnalysis
MESH = None             # MediaPipe FaceMesh helper

# quality thresholds
BLUR_MIN = 20.0
ILLUM_MIN = 30.0
ILLUM_MAX = 220.0

STABLE_NEEDED = 6          # frames of stable pose before capture
COOLDOWN_FRAMES = 10       # frames to wait after a capture


# =========================
# FastAPI app
# =========================

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],      # relax later for production
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# Pydantic models
# =========================

class RegisterResponse(BaseModel):
    success: bool
    message: str


class RecognisedStudent(BaseModel):
    student_id: str
    name: str
    best_similarity: float


class AttendanceResponse(BaseModel):
    recognised: List[RecognisedStudent]


class DebugState(BaseModel):
    num_students: int
    num_embeddings: int
    embedding_dim: int

class AttendanceRow(BaseModel):
    date: int
    student_id: str
    name: str
    time: str


class AttendanceLogResponse(BaseModel):
    year: int
    branch: str
    month: str
    rows: List[AttendanceRow]



# =========================
# InsightFace init (buffalo_l)
# =========================

def init_face_app():
    global APP_MODEL
    if APP_MODEL is not None:
        return APP_MODEL

    try:
        from insightface.app import FaceAnalysis
    except ImportError:
        print("ERROR: insightface not installed. Run: pip install insightface onnxruntime opencv-python")
        sys.exit(1)

    providers = ["CPUExecutionProvider"]
    try:
        import onnxruntime as ort
        providers = ort.get_available_providers() or ["CPUExecutionProvider"]
    except Exception:
        pass

    ctx_id = 0 if "CUDAExecutionProvider" in providers else -1

    print("[INFO] Loading InsightFace model 'buffalo_l' …")
    app_model = FaceAnalysis(name="buffalo_l", providers=providers)
    app_model.prepare(ctx_id=ctx_id, det_size=(640, 640))
    print("[INFO] InsightFace model loaded.")
    APP_MODEL = app_model
    return APP_MODEL


# =========================
# MediaPipe FaceMesh helper
# =========================

def init_mesh():
    global MESH
    if MESH is not None:
        return MESH

    try:
        import mediapipe as mp
    except ImportError:
        print("ERROR: mediapipe not installed. Run: pip install mediapipe")
        sys.exit(1)

    mesh = mp.solutions.face_mesh.FaceMesh(
        static_image_mode=False,
        max_num_faces=1,
        refine_landmarks=True,
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5,
    )
    drawing = mp.solutions.drawing_utils
    style = mp.solutions.drawing_styles
    MESH = (mesh, drawing, style, mp)
    return MESH


# =========================
# FS helpers
# =========================

def ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EMB_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    ATTENDANCE_BASE.mkdir(parents=True, exist_ok=True)   # NEW


def load_students() -> Dict[str, str]:
    if not STUDENTS_JSON.exists():
        return {}
    try:
        with open(STUDENTS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {str(k): str(v) for k, v in data.items()}
        return {}
    except Exception as e:
        print(f"[WARN] Failed to read {STUDENTS_JSON}: {e}")
        return {}


def save_students(students: Dict[str, str]):
    with open(STUDENTS_JSON, "w", encoding="utf-8") as f:
        json.dump(students, f, indent=2, ensure_ascii=False)


# =========================
# Quality + yaw helpers
# =========================

def calc_blur(image_bgr: np.ndarray) -> float:
    return float(cv2.Laplacian(image_bgr, cv2.CV_64F).var())


def calc_illum(image_bgr: np.ndarray) -> float:
    g = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
    return float(g.mean())


# mediapipe landmark indices
LM_NOSE_TIP = 1
LM_CHIN     = 152
LM_EYE_L    = 33
LM_EYE_R    = 263
LM_MOUTH_L  = 61
LM_MOUTH_R  = 291

MODEL_POINTS_3D = np.array([
    [0.0,   0.0,   0.0],     # Nose tip
    [0.0, -63.6, -12.5],     # Chin
    [-43.3, 32.7, -26.0],    # Left eye outer corner
    [ 43.3, 32.7, -26.0],    # Right eye outer corner
    [-28.9,-28.9, -24.1],    # Left mouth corner
    [ 28.9,-28.9, -24.1],    # Right mouth corner
], dtype=np.float64)


def estimate_yaw_deg(face_landmarks, w: int, h: int) -> Optional[float]:
    pts2d = []
    for idx in [LM_NOSE_TIP, LM_CHIN, LM_EYE_L, LM_EYE_R, LM_MOUTH_L, LM_MOUTH_R]:
        lm = face_landmarks.landmark[idx]
        pts2d.append([lm.x * w, lm.y * h])
    image_points = np.array(pts2d, dtype=np.float64)

    f = w
    cx, cy = w / 2.0, h / 2.0
    cam_mtx = np.array([[f, 0, cx],
                        [0, f, cy],
                        [0, 0, 1]], dtype=np.float64)
    dist = np.zeros((4, 1), dtype=np.float64)

    ok, rvec, tvec = cv2.solvePnP(MODEL_POINTS_3D, image_points, cam_mtx, dist,
                                  flags=cv2.SOLVEPNP_ITERATIVE)
    if not ok:
        return None

    R, _ = cv2.Rodrigues(rvec)
    sy = math.sqrt(R[0, 0] ** 2 + R[1, 0] ** 2)
    yaw = math.degrees(math.atan2(R[2, 0], sy))
    return float(yaw)


PHASE_BOUNDS = {
    "FRONT": (-12.0,  12.0),
    "LEFT":  (-40.0, -12.0),
    "RIGHT": ( 12.0,  40.0),
}


def classify_phase(yaw: Optional[float]) -> Optional[str]:
    if yaw is None:
        return None
    if PHASE_BOUNDS["FRONT"][0] <= yaw <= PHASE_BOUNDS["FRONT"][1]:
        return "FRONT"
    if PHASE_BOUNDS["LEFT"][0] <= yaw < PHASE_BOUNDS["LEFT"][1]:
        return "LEFT"
    if PHASE_BOUNDS["RIGHT"][0] < yaw <= PHASE_BOUNDS["RIGHT"][1]:
        return "RIGHT"
    return None


def draw_text(img, text, org, scale=0.7, color=(0, 255, 0), thick=2):
    cv2.putText(img, text, org, cv2.FONT_HERSHEY_SIMPLEX, scale, color, thick, cv2.LINE_AA)


def draw_progress(img, x, y, w, h, frac, ok):
    frac = max(0.0, min(1.0, float(frac)))
    cv2.rectangle(img, (x, y), (x + w, y + h), (80, 80, 80), 2)
    fill_w = int(w * frac)
    color = (0, 200, 0) if ok else (0, 165, 255)
    cv2.rectangle(img, (x + 2, y + 2), (x + 2 + fill_w, y + h - 2), color, -1)


# =========================
# Embeddings
# =========================

def extract_embeddings_from_frame(frame_bgr: np.ndarray) -> List[np.ndarray]:
    app_model = init_face_app()
    rgb = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2RGB)
    faces = app_model.get(rgb)
    embs: List[np.ndarray] = []

    for f in faces:
        if hasattr(f, "normed_embedding") and f.normed_embedding is not None:
            emb = np.asarray(f.normed_embedding, dtype=np.float32)
        else:
            e = getattr(f, "embedding", None)
            if e is None:
                continue
            v = np.asarray(e, dtype=np.float32)
            n = np.linalg.norm(v) + 1e-12
            emb = v / n
        embs.append(emb)

    return embs


def load_all_embeddings() -> Tuple[List[str], np.ndarray]:
    ensure_dirs()
    students = load_students()
    all_ids: List[str] = []
    all_embs: List[np.ndarray] = []

    for sid in students.keys():
        emb_path = EMB_DIR / f"{sid}.npy"
        if not emb_path.exists():
            continue
        try:
            arr = np.load(str(emb_path)).astype(np.float32)
        except Exception as e:
            print(f"[WARN] Could not load embeddings for {sid}: {e}")
            continue

        if arr.ndim == 1:
            arr = arr[None, :]

        norms = np.linalg.norm(arr, axis=1, keepdims=True) + 1e-12
        arr = arr / norms
        all_embs.append(arr)
        all_ids.extend([sid] * arr.shape[0])

    if not all_embs:
        return [], np.empty((0, 0), dtype=np.float32)

    embs = np.vstack(all_embs)
    return all_ids, embs


# =========================
# Guided registration (auto capture)
# =========================

def guided_capture(student_id: str, save_dir: Path) -> List[np.ndarray]:
    """
    One continuous guided capture:
      - uses FaceMesh + yaw to classify FRONT/LEFT/RIGHT
      - automatically saves frames when stable & good quality
      - stops when we have IMAGES_PER_POSE for each pose
    """
    mesh, drawing, style, mp = init_mesh()

    cap = cv2.VideoCapture(CAM_INDEX, cv2.CAP_DSHOW)
    if not cap.isOpened():
        print(f"[ERROR] Could not open camera index {CAM_INDEX}.")
        return []

    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

    window = f"Guided Registration ({student_id}) — press 'q' in this window to abort"
    cv2.namedWindow(window, cv2.WINDOW_NORMAL)

    counts = {"FRONT": 0, "LEFT": 0, "RIGHT": 0}
    embeddings: List[np.ndarray] = []

    cooldown = 0
    stable = 0
    last_phase: Optional[str] = None
    pulse = 0
    progress_w = 320

    print("\n[INFO] Guided capture started.")
    print("      Look at the camera FRONT, then slightly LEFT, then slightly RIGHT.")
    print("      The system will auto-capture 10 images per angle when your pose is stable.\n")

    total_needed = IMAGES_PER_POSE * len(counts)

    while sum(counts.values()) < total_needed:
        ok, frame = cap.read()
        if not ok:
            time.sleep(0.05)
            continue

        overlay = frame.copy()
        h, w = frame.shape[:2]

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        res = mesh.process(rgb)

        yaw = None
        phase = None
        ok_quality = False

        if res and res.multi_face_landmarks:
            fl = res.multi_face_landmarks[0]

            # draw facemesh
            drawing.draw_landmarks(
                image=overlay,
                landmark_list=fl,
                connections=mp.solutions.face_mesh.FACEMESH_TESSELATION,
                landmark_drawing_spec=None,
                connection_drawing_spec=style.get_default_face_mesh_tesselation_style(),
            )
            drawing.draw_landmarks(
                image=overlay,
                landmark_list=fl,
                connections=mp.solutions.face_mesh.FACEMESH_CONTOURS,
                landmark_drawing_spec=None,
                connection_drawing_spec=style.get_default_face_mesh_contours_style(),
            )

            xs = [int(l.x * w) for l in fl.landmark]
            ys = [int(l.y * h) for l in fl.landmark]
            x1, x2 = max(min(xs), 0), min(max(xs), w - 1)
            y1, y2 = max(min(ys), 0), min(max(ys), h - 1)
            ww, hh = x2 - x1, y2 - y1

            if ww > 160 and hh > 160:
                crop = frame[y1:y2, x1:x2]
                blur = calc_blur(crop)
                illum = calc_illum(crop)
                ok_quality = (blur >= BLUR_MIN) and (ILLUM_MIN <= illum <= ILLUM_MAX)
                draw_text(overlay, f"blur:{blur:.0f} illum:{illum:.0f}", (20, 80),
                          0.7, (255, 255, 255), 2)

            yaw = estimate_yaw_deg(fl, w, h)
            if yaw is not None:
                draw_text(overlay, f"yaw:{yaw:5.1f}°", (20, 110),
                          0.7, (255, 255, 255), 2)
                phase = classify_phase(yaw)

                if phase:
                    ymin, ymax = PHASE_BOUNDS[phase]
                    ok_angle = (ymin <= yaw <= ymax)
                    draw_text(
                        overlay,
                        f"{phase} target:{ymin:.0f}..{ymax:.0f}",
                        (20, 140),
                        0.7,
                        (0, 255, 0) if ok_angle else (0, 165, 255),
                        2,
                    )
                else:
                    draw_text(
                        overlay,
                        "Angle too extreme — turn slightly toward camera",
                        (20, 140),
                        0.7,
                        (0, 165, 255),
                        2,
                    )
        else:
            draw_text(overlay, "No face detected", (20, 80), 0.8, (0, 165, 255), 2)

        # pose stability
        if phase and ok_quality:
            if phase == last_phase:
                stable = min(stable + 1, STABLE_NEEDED)
            else:
                stable = 1
        else:
            stable = 0

        last_phase = phase

        # progress bar for stability
        draw_progress(overlay, 20, 170, progress_w, 16, stable / STABLE_NEEDED, phase is not None)

        # show counts
        text_counts = f"F={counts['FRONT']}/{IMAGES_PER_POSE}  L={counts['LEFT']}/{IMAGES_PER_POSE}  R={counts['RIGHT']}/{IMAGES_PER_POSE}"
        draw_text(overlay, text_counts, (20, h - 40), 0.8, (0, 255, 0), 2)

        # animated center frame
        pulse = (pulse + 1) % 40
        color = (0, 255, 0) if pulse < 20 else (0, 200, 255)
        cx, cy = w // 2, h // 2
        box_w, box_h = w // 3, h // 2
        cv2.rectangle(
            overlay,
            (cx - box_w // 2, cy - box_h // 2),
            (cx + box_w // 2, cy + box_h // 2),
            color,
            2,
        )

        cv2.imshow(window, overlay)
        key = cv2.waitKey(1) & 0xFF
        if key == ord("q"):
            print("[INFO] Registration aborted by user.")
            break

        if cooldown > 0:
            cooldown -= 1
            continue

        # auto-capture
        if phase and ok_quality and stable >= STABLE_NEEDED and counts[phase] < IMAGES_PER_POSE:
            ts = int(time.time() * 1000)
            fname = f"{phase.lower()}_{ts}.jpg"
            outp = save_dir / fname
            cv2.imwrite(str(outp), frame)
            print(f"[CAPTURE] {phase} -> {fname}")

            embs = extract_embeddings_from_frame(frame)
            if embs:
                embeddings.append(embs[0])
                counts[phase] += 1
            else:
                print("[WARN] No embedding found for captured frame (face too small / off).")

            cooldown = COOLDOWN_FRAMES
            stable = 0

    cap.release()
    try:
        cv2.destroyAllWindows()
    except cv2.error:
        pass

    print(f"[INFO] Capture done. Counts: {counts}")
    return embeddings


def register_student_internal(student_id: str, name: str) -> Tuple[bool, str]:
    ensure_dirs()
    students = load_students()
    students[student_id] = name
    save_students(students)

    save_dir = RAW_DIR / student_id
    save_dir.mkdir(parents=True, exist_ok=True)
    print(f"[INFO] Raw images will be saved under: {save_dir.resolve()}")

    embeddings = guided_capture(student_id, save_dir)
    if not embeddings:
        return False, "No embeddings collected. Registration aborted."

    embs_arr = np.stack(embeddings, axis=0).astype(np.float32)
    norms = np.linalg.norm(embs_arr, axis=1, keepdims=True) + 1e-12
    embs_arr = embs_arr / norms

    out_path = EMB_DIR / f"{student_id}.npy"
    np.save(str(out_path), embs_arr)

    print(f"[OK] Registration complete for {name} ({student_id}). Stored {embs_arr.shape[0]} embeddings.")
    return True, f"Registered {name} ({student_id}) with {embs_arr.shape[0]} embeddings."

def register_student_from_photos(
    student_id: str,
    name: str,
    images_bgr: List[np.ndarray],
) -> Tuple[bool, str]:
    """
    Register a student using a list of BGR images (already decoded).
    Used by the mobile /register_student_mobile endpoint.
    """
    ensure_dirs()
    students = load_students()
    students[student_id] = name
    save_students(students)

    save_dir = RAW_DIR / student_id
    save_dir.mkdir(parents=True, exist_ok=True)
    print(f"[INFO] (mobile) Raw images will be saved under: {save_dir.resolve()}")

    embeddings: List[np.ndarray] = []

    for i, img in enumerate(images_bgr):
        if img is None:
            continue

        # save the raw image for debugging
        out_path = save_dir / f"mobile_{i:02d}.jpg"
        cv2.imwrite(str(out_path), img)

        embs = extract_embeddings_from_frame(img)
        if embs:
            embeddings.append(embs[0])
        else:
            print(f"[WARN] No embedding extracted for mobile image {i}")

    if not embeddings:
        return False, "No usable faces found in uploaded photos."

    embs_arr = np.stack(embeddings, axis=0).astype(np.float32)
    norms = np.linalg.norm(embs_arr, axis=1, keepdims=True) + 1e-12
    embs_arr = embs_arr / norms

    out_path = EMB_DIR / f"{student_id}.npy"
    np.save(str(out_path), embs_arr)

    print(
        f"[OK] (mobile) Registration complete for {name} ({student_id}). "
        f"Stored {embs_arr.shape[0]} embeddings."
    )
    return True, f"Registered {name} ({student_id}) from mobile with {embs_arr.shape[0]} embeddings."


@app.post("/register_student_mobile", response_model=RegisterResponse)
async def api_register_student_mobile(
    student_id: str = Form(...),
    name: str = Form(...),
    photos: List[UploadFile] = File(...),
):
    """
    Mobile registration endpoint.
    Receives multiple photos taken on the phone, extracts embeddings,
    and stores them without using the server webcam.
    """
    print(f"\n[API] /register_student_mobile called for {student_id} ({name})")
    images_bgr: List[np.ndarray] = []

    for i, file in enumerate(photos):
        data = await file.read()
        nparr = np.frombuffer(data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None:
            print(f"[WARN] Could not decode uploaded photo index {i}")
            continue
        images_bgr.append(img)

    ok, msg = register_student_from_photos(student_id, name, images_bgr)
    return RegisterResponse(success=ok, message=msg)
# =========================
# Photo recognition internal
# =========================

def recognise_in_photo(image_bgr: np.ndarray) -> List[RecognisedStudent]:
    ensure_dirs()
    students = load_students()
    results: List[RecognisedStudent] = []

    app_model = init_face_app()
    rgb = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
    faces = app_model.get(rgb)

    if not faces:
        print("[INFO] No faces detected in this photo.")
        return results

    student_ids_db, embs_db = load_all_embeddings()
    if embs_db.size == 0:
        print("[INFO] No embeddings stored. Register at least one student first.")
        return results

    recognised_ids: Set[str] = set()
    best_sim_per_id: Dict[str, float] = {}

    for f in faces:
        if hasattr(f, "normed_embedding") and f.normed_embedding is not None:
            emb = np.asarray(f.normed_embedding, dtype=np.float32)
        else:
            e = getattr(f, "embedding", None)
            if e is None:
                continue
            v = np.asarray(e, dtype=np.float32)
            n = np.linalg.norm(v) + 1e-12
            emb = v / n

        sims = embs_db @ emb
        idx = int(np.argmax(sims))
        best_sim = float(sims[idx])
        sid = student_ids_db[idx]

        print(f"[DEBUG] best_sim={best_sim:.3f} for candidate {sid}")

        if best_sim >= SIM_THRESHOLD:
            recognised_ids.add(sid)
            if sid not in best_sim_per_id or best_sim > best_sim_per_id[sid]:
                best_sim_per_id[sid] = best_sim

    for sid in sorted(recognised_ids):
        name = students.get(sid, "(unknown)")
        results.append(RecognisedStudent(
            student_id=sid,
            name=name,
            best_similarity=best_sim_per_id.get(sid, 0.0),
        ))

    return results

VALID_YEARS = {1, 2, 3, 4}
VALID_BRANCHES = {
    "Artificial Intelligence",
    "Computer Science",
    "Electronics & Communication",
    "Electrical",
    "VLSI",
    "Information Technology",
}

def get_attendance_file(year: int, branch: str) -> Path:
    """
    Build folder:
      data/attendance/Year<year>/<branch>/attendance.xlsx
    and ensure the folders exist.
    """
    if year not in VALID_YEARS:
        raise ValueError(f"Invalid year: {year}")

    if branch not in VALID_BRANCHES:
        raise ValueError(f"Invalid branch: {branch}")

    folder = ATTENDANCE_BASE / f"Year{year}" / branch
    folder.mkdir(parents=True, exist_ok=True)
    return folder / "attendance.xlsx"



def write_attendance_to_excel(
    recognised: List[RecognisedStudent],
    year: int,
    branch: str,
) -> None:
    """
    Append rows to the correct Excel file:
      data/attendance/Year<year>/<branch>/attendance.xlsx

    Each workbook has one sheet per month (e.g. "November"),
    with columns: Date | Student ID | Name | Time
    """
    if not recognised:
        return

    if Workbook is None or load_workbook is None:
        print("[WARN] openpyxl not installed, skipping Excel logging. Run: pip install openpyxl")
        return

    ensure_dirs()
    attendance_file = get_attendance_file(year, branch)

    now = datetime.now()
    month_name = now.strftime("%B")       # e.g. "November"
    time_str = now.strftime("%I:%M %p")   # e.g. "09:00 AM"

    if attendance_file.exists():
        wb = load_workbook(attendance_file)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    if month_name in wb.sheetnames:
        ws = wb[month_name]
    else:
        ws = wb.create_sheet(title=month_name)
        ws.append(["Date", "Student ID", "Name", "Time"])

    for student in recognised:
        ws.append([now.day, student.student_id, student.name, time_str])

    try:
        wb.save(attendance_file)
        print(
            f"[OK] Wrote {len(recognised)} attendance rows to "
            f"{attendance_file} (sheet={month_name})."
        )
    except PermissionError as e:
        print(
            f"[ERROR] Could not save attendance Excel file "
            f"(is it open in Excel?). Skipping save. Details: {e}"
        )




# =========================
# FastAPI endpoints
# =========================

@app.on_event("startup")
def _startup():
    ensure_dirs()


@app.get("/debug_state", response_model=DebugState)
def debug_state():
    students = load_students()
    student_ids_db, embs_db = load_all_embeddings()
    if embs_db.size == 0:
        emb_dim = 0
    else:
        emb_dim = int(embs_db.shape[1])

    return DebugState(
        num_students=len(students),
        num_embeddings=embs_db.shape[0],
        embedding_dim=emb_dim,
    )


@app.post("/register_student", response_model=RegisterResponse)
def api_register_student(
    student_id: str = Form(...),
    name: str = Form(...),
):
    """
    Start guided webcam registration for a student.
    NOTE: This will open a camera window on the server machine.
          The HTTP request returns when capture finishes.
    """
    print(f"\n[API] /register_student called for {student_id} ({name})")
    ok, msg = register_student_internal(student_id, name)
    return RegisterResponse(success=ok, message=msg)


@app.post("/mark_attendance", response_model=AttendanceResponse)
async def api_mark_attendance(
    year: int = Form(...),
    branch: str = Form(...),
    photo: UploadFile = File(...),
):
    """
    Upload a class/group photo and get which registered students are present.
    Attendance is saved in:
      data/attendance/Year<year>/<branch>/attendance.xlsx
    """
    print(
        f"\n[API] /mark_attendance for year={year}, "
        f"branch={branch}, file={photo.filename}"
    )

    data = await photo.read()
    nparr = np.frombuffer(data, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None:
        print("[ERROR] Could not decode uploaded image.")
        return AttendanceResponse(recognised=[])

    recognised = recognise_in_photo(img)

    try:
        write_attendance_to_excel(recognised, year, branch)
    except ValueError as ve:
        print(f"[ERROR] {ve}")  # invalid year/branch

    return AttendanceResponse(recognised=recognised)
